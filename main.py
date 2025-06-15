from datetime import datetime
import openpyxl
import requests
import time
import json
import smtplib
from email.message import EmailMessage
import os

config = {}
api_key = ""
symbol = ""
excel_file = ""


def load_config():
    with open("config.json") as file:
        global config
        config = json.load(file)
        global api_key, symbol, excel_file
        api_key = config["api_key"]
        symbol = config["symbol"]
        excel_file = config["excel_file"]


def append_price(timestamp, symbol, price):
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        sheet.append([timestamp, symbol, price])
        workbook.save(excel_file)
        print(f"{timestamp} | {symbol} | ₹{price} added to excel")
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["Timestamp", "Symbol", "Price"])
        workbook.save(excel_file)
        sheet.append([timestamp, symbol, price])
        workbook.save(excel_file)
        print(f"{timestamp} | {symbol} | ₹{price} added to excel")


def clear_sheet_data():
    if os.path.exists(excel_file):
        print("Clearing data")
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        sheet.delete_rows(2, sheet.max_row)
        workbook.save(excel_file)


def fetch_stock_price():
    url = f"https://api.twelvedata.com/price?symbol={symbol}&apikey={api_key}"
    response = requests.get(url)
    data = response.json()
    if response.status_code == 200:
        return data['price']


def run():
    market_close_time = datetime.strptime(config["market_close_time"], "%H:%M").time()
    while True:
        now = datetime.now().time()
        if now >= market_close_time:
            print("Market is closed !!")
            break
        price = fetch_stock_price()
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if price:
            append_price(timestamp, symbol, price)
        else:
            print("Failed to fetch price.")

        time.sleep(config["interval"])


def send_email_with_attachment(file_path, sender_email, sender_password, receiver_email):
    msg = EmailMessage()
    msg["Subject"] = "Today's Stock Prices 📈"
    msg["From"] = sender_email
    msg["To"] = receiver_email
    msg.set_content("Hi,\n\nPlease find attached the stock prices for today.\n\nRegards,\nYour Stock Tracker Bot")

    # Attach Excel file
    with open(file_path, "rb") as f:
        file_data = f.read()
        file_name = f.name

    msg.add_attachment(file_data, maintype="application", subtype="vnd.openxmlformats-officedocument.spreadsheetml"
                                                                  ".sheet", filename=file_name)

    # Send email via SMTP
    with smtplib.SMTP("smtp.gmail.com", 587) as smtp:
        smtp.starttls()
        smtp.login(sender_email, sender_password)
        smtp.send_message(msg)

    print("✅ Email sent successfully!")


load_config()
clear_sheet_data()
run()
send_email_with_attachment(
    excel_file,
    sender_email=config["sender_email"],
    sender_password=config["password"],
    receiver_email=config["receiver_email"]
)